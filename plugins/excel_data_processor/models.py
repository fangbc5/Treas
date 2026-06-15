"""数据模型定义"""

from dataclasses import dataclass, field
from typing import Optional, List


@dataclass
class FieldConfig:
    """字段配置（对应模板的一个表头列）"""
    name: str
    value_type: str = "text"       # text / integer / float
    precision: int = 2             # float 时的精度
    formula: str = ""              # 取值逻辑（表达式）
    is_seed: bool = False          # 种子字段（值来自外部列表）
    enabled: bool = True           # 启用/禁用（删除）


@dataclass
class PivotConfig:
    """数据透视表配置"""
    row_field: str = ""
    value_field: str = ""
    agg_func: str = "sum"          # sum/count/avg/max/min
    filter_field: str = ""
    filter_value: str = ""


@dataclass
class DataSourceConfig:
    """数据源配置"""
    name: str
    file_path: str = ""
    sheet_name: str = ""
    header_row: int = 1
    is_pivot: bool = False
    pivot: PivotConfig = field(default_factory=PivotConfig)
    _columns: List[str] = field(default_factory=list)
    _rows: list = field(default_factory=list)

    def load(self):
        from openpyxl import load_workbook
        # 先用 read_only=True（内存优），若读取异常（某些 xlsx 兼容性问题导致
        # 只读到极少行列）则回退到 read_only=False 确保数据完整
        rows = self._read_rows(read_only=True)
        if not rows or len(rows) <= 1 or (rows and len(rows[0]) <= 1):
            rows = self._read_rows(read_only=False)
        if not rows:
            return
        header_idx = self.header_row - 1
        if header_idx >= len(rows):
            return
        self._columns = [str(c) if c is not None else "" for c in rows[header_idx]]
        data_rows = []
        for row in rows[header_idx + 1:]:
            if all(v is None for v in row):
                continue
            record = {}
            for i, col in enumerate(self._columns):
                record[col] = row[i] if i < len(row) else None
            data_rows.append(record)
        self._rows = data_rows
        if self.is_pivot:
            self._build_pivot()

    def _read_rows(self, read_only: bool):
        """读取工作表所有行（values_only）"""
        from openpyxl import load_workbook
        try:
            wb = load_workbook(self.file_path, data_only=True, read_only=read_only)
            ws = wb[self.sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
            return rows
        except Exception:
            return []

    def _build_pivot(self):
        pivot_map = {}
        for row in self._rows:
            key = row.get(self.pivot.row_field)
            if key is None:
                continue
            if self.pivot.filter_field and self.pivot.filter_value:
                fv = row.get(self.pivot.filter_field)
                if str(fv) != str(self.pivot.filter_value):
                    continue
            val = row.get(self.pivot.value_field)
            val_num = _to_number(val)
            pivot_map.setdefault(key, []).append(val_num)
        agg = self.pivot.agg_func.lower()
        result_rows = []
        for key, vals in pivot_map.items():
            vv = [v for v in vals if v is not None]
            if agg == "count":
                av = len(vv)
            elif agg == "avg":
                av = sum(vv) / len(vv) if vv else 0
            elif agg == "max":
                av = max(vv) if vv else None
            elif agg == "min":
                av = min(vv) if vv else None
            else:
                av = sum(vv) if vv else 0
            result_rows.append({self.pivot.row_field: key, self.pivot.value_field: av})
        self._columns = [self.pivot.row_field, self.pivot.value_field]
        self._rows = result_rows

    @property
    def columns(self):
        return self._columns

    @property
    def rows(self):
        return self._rows

    def find_column(self, col_letter):
        if len(col_letter) <= 3 and col_letter.isalpha():
            idx = _col_letter_to_index(col_letter)
            if 0 <= idx < len(self._columns):
                return self._columns[idx]
        if col_letter in self._columns:
            return col_letter
        return None


@dataclass
class RowContext:
    values: dict = field(default_factory=dict)

    def get(self, name):
        return self.values.get(name)

    def set(self, name, value):
        self.values[name] = value


def _to_number(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return val
    try:
        s = str(val).strip().replace(",", "").replace("%", "")
        if "." in s:
            return float(s)
        return int(s)
    except (ValueError, TypeError):
        return None


def _col_letter_to_index(letters):
    result = 0
    for ch in letters.upper():
        result = result * 26 + (ord(ch) - ord("A") + 1)
    return result - 1
