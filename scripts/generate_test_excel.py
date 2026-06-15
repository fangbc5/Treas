#!/usr/bin/env python3
"""为 excel_data_processor 插件生成测试 Excel 数据

生成文件（输出到 plugins/excel_data_processor/test_data/）：
  数据源：
    1. 员工信息.xlsx   - 测试 VLOOKUP
    2. 销售明细.xlsx   - 测试 SUMIF / COUNTIF / AVERAGEIF / MAXIFS / MINIFS
  模板：
    3. 薪资汇总模板.xlsx - 测试 VLOOKUP + 算术 + IF + ROUND
    4. 销售统计模板.xlsx - 测试聚合函数

用法:
    python3 scripts/generate_test_excel.py
"""

import os
import random
from openpyxl import Workbook

# 输出目录
OUTPUT_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
    "plugins", "excel_data_processor", "test_data",
)


# ─────────────────────────── 数据源 1：员工信息 ───────────────────────────

def create_employee_info(path: str):
    """员工信息表（VLOOKUP 数据源）"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 表头
    headers = ["员工ID", "姓名", "部门", "职位", "基本工资", "绩效系数"]
    ws.append(headers)

    # 数据（8 名员工，覆盖 4 个部门）
    employees = [
        ("E001", "张三", "技术部", "高级工程师", 15000, 1.20),
        ("E002", "李四", "技术部", "工程师",     11000, 1.10),
        ("E003", "王五", "销售部", "销售经理",   12000, 1.30),
        ("E004", "赵六", "销售部", "销售代表",    9000, 1.15),
        ("E005", "钱七", "人事部", "HR主管",     13000, 1.10),
        ("E006", "孙八", "人事部", "HR专员",      8000, 1.05),
        ("E007", "周九", "财务部", "财务经理",   14000, 1.15),
        ("E008", "吴十", "财务部", "会计",       10000, 1.08),
    ]
    for emp in employees:
        ws.append(emp)

    # 调整列宽，便于查看
    for col_letter in ["A", "B", "C", "D", "E", "F"]:
        ws.column_dimensions[col_letter].width = 12

    wb.save(path)
    print(f"✅ 生成数据源: {path}  ({len(employees)} 条记录)")


# ─────────────────────────── 数据源 2：销售明细 ───────────────────────────

def create_sales_detail(path: str):
    """销售明细表（SUMIF / COUNTIF / MAXIFS 等数据源）"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # 表头
    headers = ["订单号", "销售员", "产品", "数量", "单价", "金额", "日期"]
    ws.append(headers)

    # 4 名销售员（对应员工表中的销售部 + 2 名跨部门兼职）
    sales_people = ["李四", "王五", "赵六", "张三"]
    products = ["键盘", "鼠标", "显示器", "主机", "耳机", "摄像头"]

    random.seed(42)  # 固定随机种子，保证可复现
    rows = []
    order_no = 1001
    for _ in range(36):
        person = random.choice(sales_people)
        product = random.choice(products)
        qty = random.randint(1, 20)
        price = random.choice([99, 199, 299, 599, 899, 1299])
        amount = qty * price
        month = random.randint(1, 6)
        day = random.randint(1, 28)
        date_str = f"2024-{month:02d}-{day:02d}"
        rows.append((f"ORD{order_no}", person, product, qty, price, amount, date_str))
        order_no += 1

    for row in rows:
        ws.append(row)

    # 列宽
    for col_letter, width in zip(["A", "B", "C", "D", "E", "F", "G"],
                                 [12, 10, 12, 8, 10, 12, 14]):
        ws.column_dimensions[col_letter].width = width

    wb.save(path)
    print(f"✅ 生成数据源: {path}  ({len(rows)} 条记录)")


# ─────────────────────────── 模板 1：薪资汇总 ───────────────────────────

def create_salary_template(path: str):
    """薪资汇总模板（仅表头，数据由公式生成）

    种子字段：员工ID（从员工信息表提取唯一值）
    字段公式:
        姓名:     VLOOKUP({员工ID}, 员工信息!A:F, 2)
        部门:     VLOOKUP({员工ID}, 员工信息!A:F, 3)
        基本工资:  VLOOKUP({员工ID}, 员工信息!A:F, 5)
        绩效系数:  VLOOKUP({员工ID}, 员工信息!A:F, 6)
        绩效工资:  {基本工资} * ({绩效系数} - 1)
        应发工资:  {基本工资} + {绩效工资}
        个税:     IF({应发工资} > 15000, ROUND({应发工资} * 0.2, 2), ROUND({应发工资} * 0.1, 2))
        实发工资:  {应发工资} - {个税}
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "薪资汇总"

    headers = ["员工ID", "姓名", "部门", "基本工资", "绩效系数",
               "绩效工资", "应发工资", "个税", "实发工资"]
    ws.append(headers)

    for col_letter in ["A", "B", "C", "D", "E", "F", "G", "H", "I"]:
        ws.column_dimensions[col_letter].width = 12

    wb.save(path)
    print(f"✅ 生成模板:   {path}  ({len(headers)} 列)")


# ─────────────────────────── 模板 2：销售统计 ───────────────────────────

def create_sales_stat_template(path: str):
    """销售统计模板（仅表头，数据由公式生成）

    种子字段：销售员
    字段公式:
        总销售额:   SUMIF(销售明细!B:B, {销售员}, 销售明细!F:F)
        订单数:     COUNTIF(销售明细!B:B, {销售员})
        平均单值:   AVERAGEIF(销售明细!B:B, {销售员}, 销售明细!F:F)
        最高单值:   MAXIFS(销售明细!F:F, 销售明细!B:B, {销售员})
        最低单值:   MINIFS(销售明细!F:F, 销售明细!B:B, {销售员})
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "销售统计"

    headers = ["销售员", "总销售额", "订单数", "平均单值", "最高单值", "最低单值"]
    ws.append(headers)

    for col_letter in ["A", "B", "C", "D", "E", "F"]:
        ws.column_dimensions[col_letter].width = 12

    wb.save(path)
    print(f"✅ 生成模板:   {path}  ({len(headers)} 列)")


# ─────────────────────────── README 配置说明 ───────────────────────────

def create_readme(path: str):
    """生成配置说明文档"""
    content = """# Excel 数据处理器 - 测试数据

本目录包含用于测试 `excel_data_processor` 插件的示例数据。

## 文件说明

| 文件 | 类型 | 说明 |
|------|------|------|
| `员工信息.xlsx` | 数据源 | 8 名员工的基础信息（VLOOKUP 数据源） |
| `销售明细.xlsx` | 数据源 | 36 条销售订单（SUMIF/COUNTIF/MAXIFS 数据源） |
| `薪资汇总模板.xlsx` | 模板 | 薪资计算输出模板（VLOOKUP + 算术 + IF） |
| `销售统计模板.xlsx` | 模板 | 销售统计输出模板（聚合函数） |

## 使用步骤

### 场景一：薪资汇总（测试 VLOOKUP + 算术 + IF + ROUND）

1. **数据源标签页**：添加数据源 `员工信息`，指向 `员工信息.xlsx` 的 `Sheet1`
2. **模板标签页**：加载模板 `薪资汇总模板.xlsx`
3. **字段标签页**：按下表配置字段（**按顺序**，因为后置字段依赖前置字段）：

| 顺序 | 字段名 | 取值类型 | 精度 | 是否种子 | 公式 |
|------|--------|----------|------|----------|------|
| 1 | 员工ID | text | - | ✅ | - |
| 2 | 姓名 | text | - | | `VLOOKUP({员工ID}, 员工信息!A:F, 2)` |
| 3 | 部门 | text | - | | `VLOOKUP({员工ID}, 员工信息!A:F, 3)` |
| 4 | 基本工资 | integer | - | | `VLOOKUP({员工ID}, 员工信息!A:F, 5)` |
| 5 | 绩效系数 | float | 2 | | `VLOOKUP({员工ID}, 员工信息!A:F, 6)` |
| 6 | 绩效工资 | float | 2 | | `{基本工资} * ({绩效系数} - 1)` |
| 7 | 应发工资 | float | 2 | | `{基本工资} + {绩效工资}` |
| 8 | 个税 | float | 2 | | `IF({应发工资} > 15000, ROUND({应发工资} * 0.2, 2), ROUND({应发工资} * 0.1, 2))` |
| 9 | 实发工资 | float | 2 | | `{应发工资} - {个税}` |

4. **执行标签页**：点击"开始处理"，生成结果 Excel

### 场景二：销售统计（测试 SUMIF / COUNTIF / AVERAGEIF / MAXIFS / MINIFS）

1. **数据源标签页**：添加数据源 `销售明细`，指向 `销售明细.xlsx` 的 `Sheet1`
2. **模板标签页**：加载模板 `销售统计模板.xlsx`
3. **字段标签页**：

| 顺序 | 字段名 | 取值类型 | 精度 | 是否种子 | 公式 |
|------|--------|----------|------|----------|------|
| 1 | 销售员 | text | - | ✅ | - |
| 2 | 总销售额 | integer | - | | `SUMIF(销售明细!B:B, {销售员}, 销售明细!F:F)` |
| 3 | 订单数 | integer | - | | `COUNTIF(销售明细!B:B, {销售员})` |
| 4 | 平均单值 | float | 2 | | `AVERAGEIF(销售明细!B:B, {销售员}, 销售明细!F:F)` |
| 5 | 最高单值 | integer | - | | `MAXIFS(销售明细!F:F, 销售明细!B:B, {销售员})` |
| 6 | 最低单值 | integer | - | | `MINIFS(销售明细!F:F, 销售明细!B:B, {销售员})` |

4. **执行标签页**：点击"开始处理"

## 预期结果

### 薪资汇总（部分示例）
- E001 张三（技术部）：基本工资 15000，绩效系数 1.20，应发 18000，个税 3600，实发 14400
- E006 孙八（人事部）：基本工资 8000，绩效系数 1.05，应发 8400，个税 840，实发 7560

### 销售统计
- 4 名销售员各自的总额、订单数、平均值、最高/最低单值（因使用固定随机种子 seed=42，结果可复现）

---
由 `scripts/generate_test_excel.py` 生成，可随时重新生成。
"""
    with open(path, "w", encoding="utf-8") as f:
        f.write(content)
    print(f"✅ 生成文档:   {path}")


# ─────────────────────────── 主函数 ───────────────────────────

def main():
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    print(f"输出目录: {OUTPUT_DIR}\n")

    # 数据源
    create_employee_info(os.path.join(OUTPUT_DIR, "员工信息.xlsx"))
    create_sales_detail(os.path.join(OUTPUT_DIR, "销售明细.xlsx"))

    # 模板
    create_salary_template(os.path.join(OUTPUT_DIR, "薪资汇总模板.xlsx"))
    create_sales_stat_template(os.path.join(OUTPUT_DIR, "销售统计模板.xlsx"))

    # 说明文档
    create_readme(os.path.join(OUTPUT_DIR, "README.md"))

    print(f"\n🎉 全部生成完成！共 5 个文件位于:")
    print(f"   {OUTPUT_DIR}")


if __name__ == "__main__":
    main()