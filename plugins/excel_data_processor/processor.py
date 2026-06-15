"""Excel 处理器 - 流水线执行 + 输出"""

from typing import List
from openpyxl import load_workbook
from openpyxl.cell.cell import Cell

from models import FieldConfig, DataSourceConfig, RowContext
from formula_engine import FormulaEngine, FormulaError


def read_template_headers(template_path: str) -> List[str]:
    """读取模板第一行表头"""
    wb = load_workbook(template_path, read_only=True)
    ws = wb.active
    headers = []
    for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=False)):
        if cell.value is not None:
            headers.append(str(cell.value))
        else:
            break
    wb.close()
    return headers


def extract_seed_values(ds: DataSourceConfig, field_name: str) -> list:
    """从数据源提取种子字段的唯一值列表"""
    ds.load()
    seen = []
    for row in ds.rows:
        v = row.get(field_name)
        if v is not None and v not in seen:
            seen.append(v)
    return seen


def execute_pipeline(
    fields: List[FieldConfig],
    data_sources: dict,
    seed_values: list,
    seed_field_name: str,
    on_progress=None,
):
    """执行流水线，生成所有输出行

    Args:
        fields: 按顺序排列的字段配置
        data_sources: {名称: DataSourceConfig}
        seed_values: 种子值列表
        seed_field_name: 种子字段名
        on_progress: 进度回调 (current, total)

    Returns:
        (rows, errors) - rows: [{字段名: 值}, ...], errors: [错误信息, ...]
    """
    engine = FormulaEngine(data_sources)
    result_rows = []
    errors = []
    total = len(seed_values)

    for i, seed_val in enumerate(seed_values):
        ctx = RowContext()
        ctx.set(seed_field_name, seed_val)
        ctx.set("_row_index", i)  # 注入行号(0起始)，供 SEQ 函数使用

        row_errors = []
        for field in fields:
            if not field.enabled:
                continue
            if field.is_seed:
                continue
            if not field.formula.strip():
                ctx.set(field.name, "")
                continue
            try:
                value = engine.evaluate(field.formula, ctx)
                value = _apply_type(value, field.value_type, field.precision)
                ctx.set(field.name, value)
            except FormulaError as e:
                row_errors.append(f"{field.name}: {e}")
                ctx.set(field.name, "")

        result_rows.append(dict(ctx.values))
        if row_errors:
            errors.append(f"第{i+1}行({seed_val}): {'; '.join(row_errors)}")

        if on_progress and (i % 10 == 0 or i == total - 1):
            on_progress(i + 1, total)

    return result_rows, errors


def _apply_type(value, value_type: str, precision: int):
    """应用取值类型转换"""
    if value is None or value == "":
        return ""

    if value_type == "text":
        return str(value)

    # 尝试转数字
    try:
        if isinstance(value, str):
            num = float(value.replace(",", "").replace("%", ""))
        else:
            num = float(value)
    except (ValueError, TypeError):
        # 转不了就保持原样作为文本
        return str(value)

    if value_type == "integer":
        return str(int(round(num)))
    if value_type == "float":
        fmt = f"{{:.{precision}f}}"
        return fmt.format(num)
    return str(value)


def write_output(
    template_path: str,
    output_path: str,
    fields: List[FieldConfig],
    rows: list,
):
    """将结果写入输出 Excel（所有单元格为文本格式）

    策略：复制模板，从第2行开始写入数据，每个单元格强制文本格式
    """
    import shutil
    shutil.copy2(template_path, output_path)

    wb = load_workbook(output_path)
    ws = wb.active

    enabled_fields = [f for f in fields if f.enabled]

    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, field in enumerate(enabled_fields, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = row_data.get(field.name, "")
            # 强制文本格式
            cell.value = str(val) if val is not None else ""
            cell.number_format = "@"

    wb.save(output_path)
    wb.close()
