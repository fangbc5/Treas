"""批量重命名目录插件 - 扫描子目录导出Excel，编辑后一键重命名"""

import sys
import os

# 设置项目根目录到 sys.path（必需）
_project_root = os.path.abspath(
    os.path.join(os.path.dirname(__file__), "..", "..")
)
if _project_root not in sys.path:
    sys.path.insert(0, _project_root)

from PyQt5.QtWidgets import (
    QVBoxLayout, QHBoxLayout, QFileDialog,
    QSizePolicy,
)
from PyQt5.QtCore import Qt

from qfluentwidgets import (
    PushButton, PrimaryPushButton, LineEdit,
    TextEdit, StrongBodyLabel, CaptionLabel,
    FluentIcon,
)

from src.core.plugin_base import PluginBase


class BatchRenameDirWidget(PluginBase):
    """批量重命名目录插件"""

    plugin_id = "batch_rename_dir"
    plugin_name = "batch_rename_dir"
    plugin_version = "1.0.0"
    plugin_description = "批量重命名目录：扫描子目录导出到Excel，编辑后一键重命名"
    plugin_icon = "FOLDER"

    def __init__(self, parent=None):
        super().__init__(parent)
        self._selected_dir = ""
        self._sub_dirs = []
        self._init_ui()

    def _init_ui(self):
        layout = QVBoxLayout(self)
        layout.setSpacing(12)
        layout.setContentsMargins(20, 20, 20, 20)

        # === 选择目录区域 ===
        dir_label = StrongBodyLabel("选择目录")
        layout.addWidget(dir_label)

        dir_layout = QHBoxLayout()
        self._dir_input = LineEdit()
        self._dir_input.setPlaceholderText("点击右侧按钮选择要扫描的目录...")
        self._dir_input.setReadOnly(True)
        dir_layout.addWidget(self._dir_input)

        browse_btn = PushButton("浏览")
        browse_btn.setIcon(FluentIcon.FOLDER)
        browse_btn.clicked.connect(self._browse_dir)
        dir_layout.addWidget(browse_btn)
        layout.addLayout(dir_layout)

        # === 操作按钮区域 ===
        btn_layout = QHBoxLayout()

        scan_btn = PrimaryPushButton("扫描子目录")
        scan_btn.setIcon(FluentIcon.SEARCH)
        scan_btn.clicked.connect(self._scan_dirs)
        btn_layout.addWidget(scan_btn)

        export_btn = PrimaryPushButton("导出到 Excel")
        export_btn.setIcon(FluentIcon.SAVE)
        export_btn.clicked.connect(self._export_excel)
        btn_layout.addWidget(export_btn)

        rename_btn = PrimaryPushButton("执行重命名")
        rename_btn.setIcon(FluentIcon.SYNC)
        rename_btn.clicked.connect(self._do_rename)
        btn_layout.addWidget(rename_btn)

        btn_layout.addStretch()
        layout.addLayout(btn_layout)

        # === 统计信息 ===
        self._stats_label = CaptionLabel("")
        self._stats_label.setStyleSheet("color: #888; font-size: 12px;")
        layout.addWidget(self._stats_label)

        # === 日志输出区域 ===
        log_label = StrongBodyLabel("操作日志")
        layout.addWidget(log_label)

        self._log_output = TextEdit()
        self._log_output.setReadOnly(True)
        self._log_output.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        self._log_output.setStyleSheet("""
            TextEdit {
                background: #1e1e1e;
                color: #d4d4d4;
                font-family: "Menlo", "Courier New", monospace;
                font-size: 12px;
                border: 1px solid #333;
                border-radius: 6px;
                padding: 8px;
            }
        """)
        layout.addWidget(self._log_output)

        self._log("欢迎使用批量重命名目录工具")
        self._log("操作步骤：选择目录 → 扫描子目录 → 导出到Excel → 编辑Excel → 执行重命名")

    def _log(self, message: str):
        """追加日志信息"""
        self._log_output.append(message)
        cursor = self._log_output.textCursor()
        cursor.movePosition(cursor.End)
        self._log_output.setTextCursor(cursor)

    def _browse_dir(self):
        """选择目录"""
        dir_path = QFileDialog.getExistingDirectory(
            self, "选择要扫描的目录", "",
            QFileDialog.ShowDirsOnly | QFileDialog.DontUseNativeDialog,
        )
        if dir_path:
            self._selected_dir = dir_path
            self._dir_input.setText(dir_path)
            self._sub_dirs.clear()
            self._stats_label.setText("")
            self._log(f"已选择目录: {dir_path}")

    def _scan_dirs(self):
        """递归扫描所有子目录"""
        if not self._selected_dir:
            self._log("⚠️ 请先选择一个目录")
            return

        if not os.path.isdir(self._selected_dir):
            self._log(f"⚠️ 目录不存在: {self._selected_dir}")
            return

        self._log(f"正在扫描: {self._selected_dir}")
        self._sub_dirs.clear()

        for root, dirs, files in os.walk(self._selected_dir):
            for d in dirs:
                full_path = os.path.join(root, d)
                rel_path = os.path.relpath(full_path, self._selected_dir)
                self._sub_dirs.append(rel_path)

        self._sub_dirs.sort()

        self._log(f"扫描完成，共找到 {len(self._sub_dirs)} 个子目录:")
        for rel_path in self._sub_dirs:
            display_path = rel_path.replace(os.sep, "/")
            self._log(f"  {display_path}")

        self._stats_label.setText(f"共扫描到 {len(self._sub_dirs)} 个子目录")

    def _export_excel(self):
        """导出子目录列表到 Excel"""
        if not self._sub_dirs:
            self._log("⚠️ 请先扫描子目录（点击「扫描子目录」按钮）")
            return

        default_name = f"rename_dirs_{os.path.basename(self._selected_dir)}.xlsx"
        save_path, _ = QFileDialog.getSaveFileName(
            self, "保存 Excel 文件", default_name,
            "Excel 文件 (*.xlsx)",
            options=QFileDialog.DontUseNativeDialog,
        )
        if not save_path:
            return

        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

            wb = Workbook()
            ws = wb.active
            ws.title = "目录重命名"

            # 表头样式
            header_fill = PatternFill(start_color="4A90D9", end_color="4A90D9", fill_type="solid")
            from openpyxl.styles import Font as XlFont
            header_font = XlFont(bold=True, size=12, color="FFFFFF")
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            ws["A1"] = "原文件夹名称"
            ws["B1"] = "新文件夹名称"
            for cell in [ws["A1"], ws["B1"]]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

            for i, rel_path in enumerate(self._sub_dirs, start=2):
                display_path = rel_path.replace(os.sep, "/")
                ws.cell(row=i, column=1, value=display_path).border = thin_border
                ws.cell(row=i, column=2, value="").border = thin_border

            ws.column_dimensions["A"].width = 50
            ws.column_dimensions["B"].width = 50

            wb.save(save_path)

            self._log(f"✅ 已导出到: {save_path}")
            self._log(f"   共 {len(self._sub_dirs)} 个子目录")
            self._log("请在 Excel 中编辑「新文件夹名称」列，保存后点击「执行重命名」")

            # 自动打开文件
            import subprocess
            if sys.platform == "darwin":
                subprocess.Popen(["open", save_path])
            elif sys.platform == "win32":
                os.startfile(save_path)
            else:
                subprocess.Popen(["xdg-open", save_path])

        except Exception as e:
            self._log(f"❌ 导出失败: {e}")

    def _do_rename(self):
        """读取 Excel 执行重命名"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, "选择 Excel 文件", "",
            "Excel 文件 (*.xlsx)",
            options=QFileDialog.DontUseNativeDialog,
        )
        if not file_path:
            return

        try:
            from openpyxl import load_workbook

            wb = load_workbook(file_path)
            ws = wb.active

            rename_pairs = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or len(row) < 2:
                    continue

                original = str(row[0]).strip() if row[0] else ""
                new_name = str(row[1]).strip() if row[1] else ""

                if not original:
                    continue

                # 统一路径分隔符
                original = original.replace("/", os.sep)
                new_name = new_name.replace("/", os.sep)

                # 跳过：新名称为空 或 与原名称完全一致（完整路径对比）
                if not new_name or new_name == original:
                    continue

                rename_pairs.append((original, new_name))

            if not rename_pairs:
                self._log("⚠️ 未找到需要重命名的目录（新名称为空或与原名称一致）")
                return

            # 按路径深度降序排列（先重命名最深层的目录）
            rename_pairs.sort(key=lambda p: p[0].count(os.sep), reverse=True)

            self._log(f"即将重命名 {len(rename_pairs)} 个目录:")
            for orig, new in rename_pairs:
                self._log(f"  {orig}  →  {new}")
            self._log("")

            success_count = 0
            fail_count = 0

            for rel_path, new_name in rename_pairs:
                full_path = os.path.join(self._selected_dir, rel_path)
                # new_name 是完整相对路径（如 rocketmqlogs4/other_days1）
                new_full_path = os.path.join(self._selected_dir, new_name)

                if not os.path.isdir(full_path):
                    self._log(f"⚠️ 跳过（目录不存在）: {rel_path}")
                    continue

                try:
                    os.rename(full_path, new_full_path)
                    self._log(f"✅ {rel_path}  →  {new_name}")
                    success_count += 1
                except OSError as e:
                    self._log(f"❌ 失败: {rel_path}  →  {new_name}  ({e})")
                    fail_count += 1

            self._log("")
            self._log(f"重命名完成！成功: {success_count}, 失败: {fail_count}")
            self._stats_label.setText(
                f"重命名完成 - 成功: {success_count}, 失败: {fail_count}"
            )

        except Exception as e:
            self._log(f"❌ 操作失败: {e}")
