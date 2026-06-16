"""Excel 数据处理插件 - 主界面"""

import sys
import os
import json
from datetime import datetime

_plugin_dir = os.path.dirname(os.path.abspath(__file__))
if _plugin_dir not in sys.path:
    sys.path.insert(0, _plugin_dir)

from PyQt5.QtWidgets import (
    QVBoxLayout, QHBoxLayout, QWidget, QFileDialog,
    QTableWidget, QTableWidgetItem, QTabWidget, QHeaderView,
    QMessageBox, QInputDialog, QLabel,
    QListWidget, QListWidgetItem, QDialog,
    QDialogButtonBox, QLineEdit, QSizePolicy,
)
from PyQt5.QtCore import Qt, QThread, pyqtSignal

from qfluentwidgets import (
    PushButton, PrimaryPushButton, CaptionLabel, TitleLabel,
    SubtitleLabel, InfoBar, InfoBarPosition, ComboBox, LineEdit,
    CheckBox, TextEdit, ProgressBar,
)

from src.core.plugin_base import PluginBase
from models import FieldConfig, DataSourceConfig
from processor import read_template_headers, execute_pipeline, write_output


class PipelineWorker(QThread):
    """后台执行流水线"""
    progress = pyqtSignal(int, int)
    finished_ok = pyqtSignal(list, list)
    error = pyqtSignal(str)

    def __init__(self, fields, data_sources, seed_values, seed_field):
        super().__init__()
        self.fields = fields
        self.data_sources = data_sources
        self.seed_values = seed_values
        self.seed_field = seed_field

    def run(self):
        try:
            rows, errors = execute_pipeline(
                self.fields, self.data_sources,
                self.seed_values, self.seed_field,
                on_progress=lambda c, t: self.progress.emit(c, t),
            )
            self.finished_ok.emit(rows, errors)
        except Exception as e:
            self.error.emit(str(e))


class ExcelDataProcessorWidget(PluginBase):
    plugin_id = "excel_data_processor"
    plugin_name = "Excel 数据处理"
    plugin_version = "1.0.0"
    plugin_description = "读取Excel模板表头，配置字段取值公式，从多个数据源按流水线填充输出"
    plugin_icon = "TABLE"

    def __init__(self, parent=None):
        super().__init__(parent)
        self._template_path = ""
        self._original_headers = []
        self._fields = []
        self._data_sources = {}
        self._seed_values = []
        self._current_config_name = ""   # 当前加载的配置名，空=新建
        self._config_snapshot = ""       # 最后一次加载/保存时的字段JSON（用于脏检测）
        self._init_ui()

    def _init_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(20, 20, 20, 20)
        layout.setSpacing(12)
        layout.addWidget(TitleLabel("📊 Excel 数据处理"))

        self._tabs = QTabWidget()
        self._tabs.setDocumentMode(True)
        self._init_template_tab()
        self._init_fields_tab()
        self._init_source_tab()
        self._init_run_tab()
        self._init_help_tab()
        layout.addWidget(self._tabs)

    # ── Tab 1: 模板 ──
    def _init_template_tab(self):
        tab = QWidget()
        lay = QVBoxLayout(tab)
        lay.setSpacing(10)

        lay.addWidget(SubtitleLabel("1️⃣ 选择输出模板"))
        row = QHBoxLayout()
        self._template_label = CaptionLabel("未选择模板文件")
        row.addWidget(self._template_label, 1)
        btn = PushButton("📂 选择模板")
        btn.clicked.connect(self._pick_template)
        row.addWidget(btn)
        lay.addLayout(row)

        self._template_headers_view = TextEdit()
        self._template_headers_view.setReadOnly(True)
        self._template_headers_view.setPlaceholderText("选择模板后，表头字段将显示在这里...")
        lay.addWidget(self._template_headers_view, 1)

        self._tabs.addTab(tab, "📋 输出模板")

    def _pick_template(self):
        path, _ = QFileDialog.getOpenFileName(
            self, "选择 Excel 模板", "", "Excel 文件 (*.xlsx *.xlsm)"
        )
        if not path:
            return
        try:
            headers = read_template_headers(path)
            self._template_path = path
            self._original_headers = headers
            self._template_label.setText(f"📄 {os.path.basename(path)} ({len(headers)} 个字段)")
            self._template_headers_view.setPlainText(
                f"模板文件: {path}\n共 {len(headers)} 个字段:\n\n" +
                "\n".join(f"  {i+1}. {h}" for i, h in enumerate(headers))
            )
            self._rebuild_fields(headers)
        except Exception as e:
            InfoBar.error("读取失败", str(e), parent=self, duration=3000, position=InfoBarPosition.TOP)

    def _rebuild_fields(self, headers):
        """从模板表头重建字段配置"""
        self._fields = []
        for i, h in enumerate(headers):
            f = FieldConfig(name=h, is_seed=(i == 0))
            self._fields.append(f)
        self._refresh_fields_table()

    # ── Tab 2: 字段配置 ──
    def _init_fields_tab(self):
        tab = QWidget()
        lay = QVBoxLayout(tab)
        lay.setSpacing(8)

        lay.addWidget(SubtitleLabel("2️⃣ 字段取值配置"))
        hint = CaptionLabel("字段按顺序执行。{字段名} 引用前序字段已计算的值。勾选「种子」列指定种子字段，种子字段公式填写「数据源名!列名」指定取值列。")
        lay.addWidget(hint)

        # 第一行：字段编辑操作
        btn_row1 = QHBoxLayout()
        reset_btn = PushButton("🔄 重置字段")
        reset_btn.clicked.connect(self._reset_fields)
        btn_row1.addWidget(reset_btn)

        del_btn = PushButton("🗑️ 删除字段")
        del_btn.clicked.connect(self._delete_field)
        btn_row1.addWidget(del_btn)

        restore_btn = PushButton("♻️ 恢复字段")
        restore_btn.clicked.connect(self._restore_field)
        btn_row1.addWidget(restore_btn)

        up_btn = PushButton("⬆️ 上移")
        up_btn.clicked.connect(lambda: self._move_field(-1))
        btn_row1.addWidget(up_btn)

        down_btn = PushButton("⬇️ 下移")
        down_btn.clicked.connect(lambda: self._move_field(1))
        btn_row1.addWidget(down_btn)

        btn_row1.addStretch()
        lay.addLayout(btn_row1)

        # 第二行：配置管理
        btn_row2 = QHBoxLayout()
        save_btn = PushButton("💾 保存配置")
        save_btn.clicked.connect(self._save_config_dialog)
        btn_row2.addWidget(save_btn)

        load_btn = PushButton("📂 加载配置")
        load_btn.clicked.connect(self._load_config_dialog)
        btn_row2.addWidget(load_btn)

        manage_btn = PushButton("🗂️ 管理配置")
        manage_btn.clicked.connect(self._manage_configs_dialog)
        btn_row2.addWidget(manage_btn)

        # 去重勾选（种子字段去重）
        self._dedup_check = CheckBox("🌱种子去重")
        self._dedup_check.setChecked(True)
        self._dedup_check.setToolTip("对种子字段提取的值去重")
        btn_row2.addWidget(self._dedup_check)

        btn_row2.addStretch()
        lay.addLayout(btn_row2)

        # 字段表格：5列（名称/类型/精度/公式/种子）
        self._fields_table = QTableWidget(0, 5)
        self._fields_table.setHorizontalHeaderLabels(
            ["字段名称", "取值类型", "精度", "取值逻辑（公式）", "种子"]
        )
        header = self._fields_table.horizontalHeader()
        # 全部使用可收缩模式，避免 ResizeToContents 撑爆宽度导致水平滚动
        header.setSectionResizeMode(0, QHeaderView.Interactive)
        self._fields_table.setColumnWidth(0, 150)
        header.setSectionResizeMode(1, QHeaderView.Interactive)
        self._fields_table.setColumnWidth(1, 100)
        header.setSectionResizeMode(2, QHeaderView.Interactive)
        self._fields_table.setColumnWidth(2, 80)
        header.setSectionResizeMode(3, QHeaderView.Stretch)
        header.setSectionResizeMode(4, QHeaderView.Interactive)
        self._fields_table.setColumnWidth(4, 60)
        self._fields_table.verticalHeader().setDefaultSectionSize(44)
        self._fields_table.verticalHeader().setMinimumSectionSize(44)
        lay.addWidget(self._fields_table, 1)

        self._tabs.addTab(tab, "⚙️ 字段配置")

    def _refresh_fields_table(self):
        table = self._fields_table
        # 先清除所有 cellWidget（安全断开信号），避免 setRowCount 缩减行数时
        # widget 析构触发 textChanged 等信号访问已变更的 _fields 导致越界崩溃
        table.clearContents()
        table.setRowCount(len(self._fields))
        for i, f in enumerate(self._fields):
            name_item = QTableWidgetItem(f.name)
            name_item.setFlags(Qt.ItemIsEnabled | Qt.ItemIsSelectable)
            # 禁用字段（已删除）应用灰色+删除线样式
            if not f.enabled:
                name_item.setForeground(Qt.gray)
                font = name_item.font()
                font.setStrikeOut(True)
                name_item.setFont(font)
            table.setItem(i, 0, name_item)

            type_combo = ComboBox()
            type_combo.addItems(["文本", "整数", "浮点数"])
            type_combo.setCurrentIndex({"text": 0, "integer": 1, "float": 2}[f.value_type])
            idx = i
            type_combo.currentIndexChanged.connect(
                lambda v, idx=idx: self._on_field_type_change(idx, v)
            )
            table.setCellWidget(i, 1, type_combo)

            prec = ComboBox()
            prec.addItems([str(n) for n in range(11)])
            prec.setCurrentIndex(min(f.precision, 10))
            prec.currentIndexChanged.connect(lambda v, idx=idx: setattr(self._fields[idx], 'precision', v))
            table.setCellWidget(i, 2, prec)

            formula_edit = LineEdit()
            # qfluentwidgets LineEdit 默认 setFixedHeight(33)，无法填满 44px 表格行高，
            # 解除固定高度限制，让控件跟随单元格行高
            formula_edit.setMinimumHeight(0)
            formula_edit.setMaximumHeight(16777215)
            formula_edit.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
            formula_edit.setText(f.formula)
            if f.is_seed:
                formula_edit.setPlaceholderText("如: 数据源名!列名")
            else:
                formula_edit.setPlaceholderText("如: VLOOKUP({项目编码}, 数据源A!A:D, 2, FALSE)")
            formula_edit.textChanged.connect(
                lambda v, idx=idx: setattr(self._fields[idx], 'formula', v)
            )
            table.setCellWidget(i, 3, formula_edit)

            seed_check = CheckBox()
            seed_check.blockSignals(True)
            seed_check.setChecked(f.is_seed)
            seed_check.blockSignals(False)
            seed_check.setStyleSheet("QCheckBox { margin-left: 20px; }")
            seed_check.toggled.connect(lambda checked, idx=idx: self._on_seed_check_change(idx, checked))
            table.setCellWidget(i, 4, seed_check)

            name_item.setTextAlignment(Qt.AlignVCenter | Qt.AlignLeft)

    def _on_seed_check_change(self, idx, checked):
        """种子勾选变更：保证唯一种子字段（不重建表格，避免清空已输入的公式）"""
        if checked:
            for i, f in enumerate(self._fields):
                f.is_seed = (i == idx)
            # 只刷新种子列勾选状态和公式 placeholder，不重建整个表格
            for i in range(self._fields_table.rowCount()):
                chk = self._fields_table.cellWidget(i, 4)
                edit = self._fields_table.cellWidget(i, 3)
                if chk:
                    chk.blockSignals(True)
                    chk.setChecked(self._fields[i].is_seed)
                    chk.blockSignals(False)
                if edit:
                    if self._fields[i].is_seed:
                        edit.setPlaceholderText("如: 数据源名!列名")
                    else:
                        edit.setPlaceholderText("如: VLOOKUP({项目编码}, 数据源A!A:D, 2, FALSE)")
        else:
            # 不允许取消唯一种子，强制恢复勾选
            self._fields[idx].is_seed = True
            chk = self._fields_table.cellWidget(idx, 4)
            if chk:
                chk.blockSignals(True)
                chk.setChecked(True)
                chk.blockSignals(False)

    def _get_seed_field_name(self) -> str:
        for f in self._fields:
            if f.is_seed:
                return f.name
        return self._fields[0].name if self._fields else ""

    def _get_seed_field_formula(self) -> str:
        for f in self._fields:
            if f.is_seed:
                return f.formula.strip()
        return ""

    def _on_field_type_change(self, idx, combo_idx):
        types = ["text", "integer", "float"]
        self._fields[idx].value_type = types[combo_idx]

    def _reset_fields(self):
        if self._original_headers:
            self._rebuild_fields(self._original_headers)

    def _delete_field(self):
        """删除字段：软删除（enabled=False），不参与计算和输出，保留列占位避免错位"""
        row = self._fields_table.currentRow()
        if row < 0:
            InfoBar.warning("提示", "请先在表格中选择要删除的字段行",
                            parent=self, duration=2000, position=InfoBarPosition.TOP)
            return
        f = self._fields[row]
        if not f.enabled:
            InfoBar.warning("提示", "该字段已删除，无需重复操作",
                            parent=self, duration=2000, position=InfoBarPosition.TOP)
            return
        if f.is_seed:
            InfoBar.warning("提示", "种子字段不可删除，请先将其他字段设为种子",
                            parent=self, duration=2500, position=InfoBarPosition.TOP)
            return
        # 至少保留一个启用的非种子字段可删（种子不可删，所以检查启用的非种子字段数）
        enabled_non_seed = sum(1 for x in self._fields if x.enabled and not x.is_seed)
        if enabled_non_seed <= 0:
            InfoBar.warning("提示", "至少需要保留一个启用字段",
                            parent=self, duration=2000, position=InfoBarPosition.TOP)
            return
        f.enabled = False
        self._refresh_fields_table()
        InfoBar.success("已删除", "该字段将不参与输出（可点击「恢复字段」还原）",
                        parent=self, duration=2500, position=InfoBarPosition.TOP)

    def _restore_field(self):
        """恢复已删除的字段（enabled=True）"""
        row = self._fields_table.currentRow()
        if row < 0:
            InfoBar.warning("提示", "请先在表格中选择要恢复的字段行",
                            parent=self, duration=2000, position=InfoBarPosition.TOP)
            return
        f = self._fields[row]
        if f.enabled:
            InfoBar.warning("提示", "该字段未被删除，无需恢复",
                            parent=self, duration=2000, position=InfoBarPosition.TOP)
            return
        f.enabled = True
        self._refresh_fields_table()
        InfoBar.success("已恢复", parent=self, duration=1500, position=InfoBarPosition.TOP)

    def _move_field(self, direction):
        row = self._fields_table.currentRow()
        if row < 0:
            InfoBar.warning("提示", "请先在表格中选择要移动的字段行",
                            parent=self, duration=2000, position=InfoBarPosition.TOP)
            return
        new_row = row + direction
        if new_row < 0 or new_row >= len(self._fields):
            return
        if row == 0 or new_row == 0:
            InfoBar.warning("提示", "种子字段必须为第一个，不可移动",
                            parent=self, duration=2000, position=InfoBarPosition.TOP)
            return
        self._fields[row], self._fields[new_row] = self._fields[new_row], self._fields[row]
        self._refresh_fields_table()
        self._fields_table.setCurrentCell(new_row, 0)

    def _sync_fields_from_table(self):
        for i, f in enumerate(self._fields):
            if i >= self._fields_table.rowCount():
                break
            type_combo = self._fields_table.cellWidget(i, 1)
            prec = self._fields_table.cellWidget(i, 2)
            formula = self._fields_table.cellWidget(i, 3)
            seed = self._fields_table.cellWidget(i, 4)
            if type_combo:
                f.value_type = ["text", "integer", "float"][type_combo.currentIndex()]
            if prec:
                f.precision = prec.currentIndex()
            if formula:
                f.formula = formula.text()
            if seed:
                f.is_seed = seed.isChecked()

    # ── 种子值提取（解析种子字段公式） ──
    def _extract_seed_values(self):
        """解析种子字段公式，提取种子值列表

        公式格式（向后兼容）:
            数据源名!列名                    -- 直接取整列原值
            数据源名!列名 >> 提取器(参数)     -- 取值后应用提取器
        提取器示例:
            SPLIT(" ", 0)   -- 按空格分割取第0段
        """
        self._sync_fields_from_table()
        formula = self._get_seed_field_formula()
        if not formula:
            self._seed_values = []
            return None, "种子字段公式为空，请填写「数据源名!列名」"

        if "!" not in formula:
            self._seed_values = []
            return None, f"种子字段公式格式应为「数据源名!列名」，当前: {formula}"

        # 分离数据源引用与可选的提取器（>> 分隔）
        extractor_fn = None
        ref_part = formula
        if ">>" in formula:
            ref_part, extractor_expr = formula.split(">>", 1)
            extractor_fn, err = self._parse_seed_extractor(extractor_expr.strip())
            if err:
                self._seed_values = []
                return None, err

        ds_name, col_name = ref_part.split("!", 1)
        ds_name = ds_name.strip()
        col_name = col_name.strip()

        ds = self._data_sources.get(ds_name)
        if not ds:
            self._seed_values = []
            return None, f"找不到数据源「{ds_name}」，请检查名称"

        try:
            ds.load()
        except Exception as e:
            self._seed_values = []
            return None, f"加载数据源「{ds_name}」失败: {e}"

        dedup = self._dedup_check.isChecked()
        vals = []
        seen = set()
        for row in ds.rows:
            v = row.get(col_name)
            if v is None:
                continue
            # 应用提取器
            if extractor_fn:
                v = extractor_fn(v)
                if v is None or str(v).strip() == "":
                    continue
            if dedup:
                key = str(v)
                if key in seen:
                    continue
                seen.add(key)
            vals.append(v)

        self._seed_values = vals
        tag = "去重后" if dedup else "全部"
        ext_tag = "（含提取器）" if extractor_fn else ""
        return vals, f"从「{ds_name}」的「{col_name}」列提取 {len(vals)} 个值（{tag}{ext_tag}）"

    def _parse_seed_extractor(self, expr):
        """解析种子提取器表达式，返回 (处理函数, 错误信息)

        支持格式: 提取器名(参数1, 参数2, ...)
        例如: SPLIT(" ", 0)  -- 按空格分割取第0段
        """
        import re
        expr = expr.strip()
        m = re.match(r'^(\w+)\s*\((.*)\)$', expr)
        if not m:
            return None, f"提取器格式错误: {expr}（应为 提取器名(参数)）"

        name = m.group(1).upper()
        args = self._parse_extractor_args(m.group(2))

        if name == "SPLIT":
            if len(args) < 2:
                return None, "SPLIT 需要2个参数: 分隔符, 索引"
            sep = args[0]
            try:
                idx = int(args[1])
            except (ValueError, TypeError):
                return None, f"SPLIT 索引必须是整数: {args[1]}"

            def split_extractor(v):
                parts = str(v).split(sep)
                if 0 <= idx < len(parts):
                    return parts[idx].strip()
                return ""
            return split_extractor, None

        return None, f"不支持的提取器: {name}"

    def _parse_extractor_args(self, args_str):
        """解析提取器参数列表（处理引号字符串和数字）"""
        args = []
        current = ""
        in_string = False
        for ch in args_str:
            if ch == '"':
                in_string = not in_string
                current += ch
            elif ch == "," and not in_string:
                args.append(self._parse_single_arg(current.strip()))
                current = ""
            else:
                current += ch
        if current.strip():
            args.append(self._parse_single_arg(current.strip()))
        return args

    @staticmethod
    def _parse_single_arg(s):
        """解析单个参数值（去引号）"""
        s = s.strip()
        if len(s) >= 2 and s.startswith('"') and s.endswith('"'):
            return s[1:-1]
        return s

    def _load_all_sources(self):
        for ds in self._data_sources.values():
            ds.load()

    # ── 字段配置存储与复用 ──

    def _get_configs_dir(self) -> str:
        from src.core.paths import get_plugin_data_dir
        path = os.path.join(get_plugin_data_dir(), self.plugin_id, "configs")
        os.makedirs(path, exist_ok=True)
        return path

    def _config_path(self, name: str) -> str:
        safe = "".join(c for c in name if c not in r'\/:*?"<>|').strip()
        if not safe:
            safe = "untitled"
        return os.path.join(self._get_configs_dir(), f"{safe}.json")

    def _serialize_fields(self) -> list:
        return [
            {
                "name": f.name,
                "value_type": f.value_type,
                "precision": f.precision,
                "formula": f.formula,
                "is_seed": f.is_seed,
                "enabled": f.enabled,
            }
            for f in self._fields
        ]

    def _save_snapshot(self):
        """记录当前字段状态的快照（用于检测后续是否有修改）"""
        self._config_snapshot = json.dumps(self._serialize_fields(), ensure_ascii=False)

    def _is_config_dirty(self) -> bool:
        """检测当前字段是否相对于上次加载/保存有改动"""
        if not self._fields:
            return False
        self._sync_fields_from_table()
        current = json.dumps(self._serialize_fields(), ensure_ascii=False)
        return current != self._config_snapshot

    def _save_config(self, name: str):
        data = {
            "version": 1,
            "saved_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "name": name,
            "fields": self._serialize_fields(),
        }
        path = self._config_path(name)
        with open(path, "w", encoding="utf-8") as fp:
            json.dump(data, fp, ensure_ascii=False, indent=2)
        return path

    def _load_config(self, name: str) -> dict:
        path = self._config_path(name)
        with open(path, "r", encoding="utf-8") as fp:
            return json.load(fp)

    def _list_configs(self) -> list:
        d = self._get_configs_dir()
        items = []
        for fn in os.listdir(d):
            if fn.endswith(".json"):
                full = os.path.join(d, fn)
                try:
                    with open(full, "r", encoding="utf-8") as fp:
                        data = json.load(fp)
                    items.append({
                        "name": data.get("name", os.path.splitext(fn)[0]),
                        "file": full,
                        "saved_at": data.get("saved_at", ""),
                        "field_count": len(data.get("fields", [])),
                        "mtime": os.path.getmtime(full),
                    })
                except Exception:
                    continue
        items.sort(key=lambda x: x.get("mtime", 0), reverse=True)
        return items

    def _delete_config(self, name: str) -> bool:
        path = self._config_path(name)
        if os.path.exists(path):
            os.remove(path)
            return True
        return False

    def _apply_config(self, data: dict):
        fields_data = data.get("fields", [])
        new_fields = []
        for i, fd in enumerate(fields_data):
            is_seed = fd.get("is_seed", i == 0)
            enabled = fd.get("enabled", True)
            new_fields.append(FieldConfig(
                name=fd.get("name", f"字段{i+1}"),
                value_type=fd.get("value_type", "text"),
                precision=fd.get("precision", 2),
                formula=fd.get("formula", ""),
                is_seed=is_seed,
                enabled=enabled,
            ))
        if not new_fields:
            return False, "配置中没有字段"
        self._fields = new_fields
        self._refresh_fields_table()
        return True, None

    def _save_config_dialog(self):
        if not self._fields:
            InfoBar.warning("提示", "当前没有字段可保存", parent=self,
                            duration=2000, position=InfoBarPosition.TOP)
            return
        self._sync_fields_from_table()

        # 已加载配置 → 直接覆盖，不弹窗
        if self._current_config_name:
            try:
                self._save_config(self._current_config_name)
                self._save_snapshot()
                InfoBar.success("配置已保存",
                                f"已更新配置 '{self._current_config_name}'",
                                parent=self, duration=2500, position=InfoBarPosition.TOP)
            except Exception as e:
                InfoBar.error("保存失败", str(e), parent=self,
                              duration=3000, position=InfoBarPosition.TOP)
            return

        # 新建 → 弹窗输入名称
        name, ok = QInputDialog.getText(self, "保存配置", "配置名称:",
                                        QLineEdit.Normal, "")
        if not ok or not name.strip():
            return
        name = name.strip()
        path = self._config_path(name)
        if os.path.exists(path):
            ret = QMessageBox.question(
                self, "确认覆盖",
                f"配置 '{name}' 已存在，是否覆盖？",
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No
            )
            if ret != QMessageBox.Yes:
                return
        try:
            self._save_config(name)
            self._current_config_name = name
            self._save_snapshot()
            InfoBar.success("保存成功",
                            f"已保存 {len(self._fields)} 个字段到 '{name}'",
                            parent=self, duration=2500, position=InfoBarPosition.TOP)
        except Exception as e:
            InfoBar.error("保存失败", str(e), parent=self,
                          duration=3000, position=InfoBarPosition.TOP)

    def _load_config_dialog(self):
        configs = self._list_configs()
        if not configs:
            InfoBar.warning("提示", "没有已保存的配置", parent=self,
                            duration=2000, position=InfoBarPosition.TOP)
            return
        dlg = QDialog(self)
        dlg.setWindowTitle("加载字段配置")
        dlg.setMinimumWidth(420)
        dlg_layout = QVBoxLayout(dlg)
        dlg_layout.addWidget(QLabel("选择要加载的配置:"))
        list_widget = QListWidget()
        for c in configs:
            display = f"{c['name']}  ({c['field_count']} 字段, {c['saved_at']})"
            item = QListWidgetItem(display)
            item.setData(Qt.UserRole, c["name"])
            list_widget.addItem(item)
        if list_widget.count() > 0:
            list_widget.setCurrentRow(0)
        dlg_layout.addWidget(list_widget)
        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(dlg.accept)
        btns.rejected.connect(dlg.reject)
        dlg_layout.addWidget(btns)
        if dlg.exec_() != QDialog.Accepted:
            return
        item = list_widget.currentItem()
        if not item:
            return
        name = item.data(Qt.UserRole)
        try:
            data = self._load_config(name)
        except Exception as e:
            InfoBar.error("加载失败", str(e), parent=self,
                          duration=3000, position=InfoBarPosition.TOP)
            return
        cfg_names = [fd.get("name", "") for fd in data.get("fields", [])]
        current_names = [f.name for f in self._fields]
        if current_names and cfg_names != current_names:
            missing = [n for n in cfg_names if n not in current_names]
            extra = [n for n in current_names if n not in cfg_names]
            msg_parts = []
            if missing:
                msg_parts.append(f"配置含当前模板没有的字段: {', '.join(missing[:5])}")
            if extra:
                msg_parts.append(f"当前模板含配置没有的字段: {', '.join(extra[:5])}")
            warn = "字段不匹配，可能已更换模板。\n" + "; ".join(msg_parts)
            warn += "\n\n字段名/数量不一致时，公式可能失效。是否继续加载？"
            ret = QMessageBox.warning(
                self, "字段不匹配", warn,
                QMessageBox.Yes | QMessageBox.No, QMessageBox.No
            )
            if ret != QMessageBox.Yes:
                return
        # 加载新配置前，检查当前配置是否有未保存修改
        if self._current_config_name and self._is_config_dirty():
            ret = QMessageBox.question(
                self, "未保存的修改",
                f"配置 '{self._current_config_name}' 已修改但未保存，是否先保存？",
                QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel,
                QMessageBox.Yes
            )
            if ret == QMessageBox.Cancel:
                return
            if ret == QMessageBox.Yes:
                try:
                    self._save_config(self._current_config_name)
                    self._save_snapshot()
                    InfoBar.success("配置已保存",
                                    f"已更新 '{self._current_config_name}'",
                                    parent=self, duration=2500, position=InfoBarPosition.TOP)
                except Exception as e:
                    InfoBar.error("保存失败", str(e), parent=self,
                                  duration=3000, position=InfoBarPosition.TOP)
                    return

        ok, err = self._apply_config(data)
        if ok:
            self._current_config_name = name
            self._save_snapshot()
            InfoBar.success("加载成功", f"已加载配置 '{name}'",
                            parent=self, duration=2500, position=InfoBarPosition.TOP)
        else:
            InfoBar.error("加载失败", err or "未知错误", parent=self,
                          duration=3000, position=InfoBarPosition.TOP)

    def _manage_configs_dialog(self):
        configs = self._list_configs()
        if not configs:
            InfoBar.warning("提示", "没有已保存的配置", parent=self,
                            duration=2000, position=InfoBarPosition.TOP)
            return
        dlg = QDialog(self)
        dlg.setWindowTitle("管理字段配置")
        dlg.setMinimumWidth(420)
        dlg_layout = QVBoxLayout(dlg)
        dlg_layout.addWidget(QLabel("选择配置并删除:"))
        list_widget = QListWidget()
        for c in configs:
            display = f"{c['name']}  ({c['field_count']} 字段, {c['saved_at']})"
            item = QListWidgetItem(display)
            item.setData(Qt.UserRole, c["name"])
            list_widget.addItem(item)
        dlg_layout.addWidget(list_widget)
        btn_row = QHBoxLayout()
        del_btn = PushButton("🗑️ 删除选中")
        del_btn.setStyleSheet("color: #d04040;")
        del_btn.clicked.connect(lambda: self._delete_selected_config(list_widget, dlg))
        btn_row.addWidget(del_btn)
        btn_row.addStretch()
        close_btn = PushButton("关闭")
        close_btn.clicked.connect(dlg.reject)
        btn_row.addWidget(close_btn)
        dlg_layout.addLayout(btn_row)
        dlg.exec_()

    def _delete_selected_config(self, list_widget, dlg):
        item = list_widget.currentItem()
        if not item:
            return
        name = item.data(Qt.UserRole)
        ret = QMessageBox.question(
            self, "确认删除",
            f"确定要删除配置 '{name}' 吗？此操作不可恢复。",
            QMessageBox.Yes | QMessageBox.No, QMessageBox.No
        )
        if ret != QMessageBox.Yes:
            return
        if self._delete_config(name):
            row = list_widget.currentRow()
            list_widget.takeItem(row)
            InfoBar.success("删除成功", f"已删除 '{name}'", parent=self,
                            duration=2000, position=InfoBarPosition.TOP)
            if list_widget.count() == 0:
                dlg.reject()
        else:
            InfoBar.error("删除失败", "文件不存在", parent=self,
                          duration=2000, position=InfoBarPosition.TOP)

    # ── Tab 3: 数据源 ──
    def _init_source_tab(self):
        tab = QWidget()
        lay = QVBoxLayout(tab)
        lay.setSpacing(8)
        lay.addWidget(SubtitleLabel("3️⃣ 数据源配置"))
        lay.addWidget(CaptionLabel("数据源名默认取文件名（如 员工信息.xlsx → 员工信息），在公式中通过名称引用（如 员工信息!A:F）"))

        btn_row = QHBoxLayout()
        add_btn = PushButton("➕ 添加数据源")
        add_btn.clicked.connect(self._add_source)
        btn_row.addWidget(add_btn)
        del_btn = PushButton("🗑️ 删除选中")
        del_btn.clicked.connect(self._del_source)
        btn_row.addWidget(del_btn)
        btn_row.addStretch()
        lay.addLayout(btn_row)

        self._source_table = QTableWidget(0, 6)
        self._source_table.setHorizontalHeaderLabels(
            ["名称", "文件", "Sheet", "透视", "透视配置", "操作"]
        )
        sh = self._source_table.horizontalHeader()
        sh.setSectionResizeMode(0, QHeaderView.Interactive)   # 名称
        self._source_table.setColumnWidth(0, 120)
        sh.setSectionResizeMode(1, QHeaderView.Stretch)        # 文件（弹性填充）
        sh.setSectionResizeMode(2, QHeaderView.Interactive)   # Sheet
        self._source_table.setColumnWidth(2, 100)
        sh.setSectionResizeMode(3, QHeaderView.Interactive)   # 透视
        self._source_table.setColumnWidth(3, 60)
        sh.setSectionResizeMode(4, QHeaderView.Stretch)        # 透视配置（弹性）
        sh.setSectionResizeMode(5, QHeaderView.Interactive)   # 操作
        self._source_table.setColumnWidth(5, 80)
        lay.addWidget(self._source_table, 1)

        self._tabs.addTab(tab, "📂 数据源")

    def _add_source(self):
        paths, _ = QFileDialog.getOpenFileNames(
            self, "选择数据源 Excel（可多选）", "", "Excel 文件 (*.xlsx *.xls)"
        )
        if not paths:
            return
        for path in paths:
            name = os.path.splitext(os.path.basename(path))[0]
            base = name
            i = 2
            while name in self._data_sources:
                name = f"{base}{i}"
                i += 1
            ds = DataSourceConfig(name=name, file_path=path, sheet_name="")
            try:
                from openpyxl import load_workbook
                wb = load_workbook(path, read_only=True)
                if wb.sheetnames:
                    ds.sheet_name = wb.sheetnames[0]
                wb.close()
            except Exception:
                pass
            self._data_sources[name] = ds
        self._refresh_source_table()
        if len(paths) > 1:
            InfoBar.success("已添加", f"已批量添加 {len(paths)} 个数据源",
                            parent=self, duration=2500, position=InfoBarPosition.TOP)

    def _del_source(self):
        row = self._source_table.currentRow()
        if row < 0:
            return
        name_item = self._source_table.item(row, 0)
        if name_item and name_item.text() in self._data_sources:
            del self._data_sources[name_item.text()]
            self._refresh_source_table()

    def _refresh_source_table(self):
        table = self._source_table
        table.setRowCount(len(self._data_sources))
        for i, (name, ds) in enumerate(self._data_sources.items()):
            table.setItem(i, 0, QTableWidgetItem(name))
            table.setItem(i, 1, QTableWidgetItem(os.path.basename(ds.file_path)))
            table.setItem(i, 2, QTableWidgetItem(ds.sheet_name))
            table.setItem(i, 3, QTableWidgetItem("✅" if ds.is_pivot else "—"))
            pv = ds.pivot
            pv_text = f"{pv.row_field}→{pv.value_field}({pv.agg_func})" if ds.is_pivot else ""
            table.setItem(i, 4, QTableWidgetItem(pv_text))
            config_btn = PushButton("配置")
            config_btn.clicked.connect(lambda _, n=name: self._config_source(n))
            table.setCellWidget(i, 5, config_btn)

    def _config_source(self, name):
        from PyQt5.QtWidgets import QDialog, QFormLayout, QDialogButtonBox
        ds = self._data_sources.get(name)
        if not ds:
            return
        dlg = QDialog(self)
        dlg.setWindowTitle(f"配置 {name}")
        dlg.setMinimumWidth(400)
        form = QFormLayout(dlg)

        sheet_combo = ComboBox()
        try:
            from openpyxl import load_workbook
            wb = load_workbook(ds.file_path, read_only=True)
            sheet_combo.addItems(wb.sheetnames)
            wb.close()
            if ds.sheet_name:
                sheet_combo.setCurrentText(ds.sheet_name)
        except Exception:
            sheet_combo.addItem(ds.sheet_name or "Sheet1")
        form.addRow("工作表:", sheet_combo)

        pivot_check = CheckBox("启用数据透视")
        pivot_check.setChecked(ds.is_pivot)
        form.addRow("透视模式:", pivot_check)

        row_edit = LineEdit()
        row_edit.setText(ds.pivot.row_field or "")
        row_edit.setPlaceholderText("行标签字段名")
        form.addRow("行标签:", row_edit)
        val_edit = LineEdit()
        val_edit.setText(ds.pivot.value_field or "")
        val_edit.setPlaceholderText("值字段名")
        form.addRow("值字段:", val_edit)
        agg_combo = ComboBox()
        agg_combo.addItems(["sum 求和", "count 计数", "avg 平均", "max 最大", "min 最小"])
        agg_map = {"sum": 0, "count": 1, "avg": 2, "max": 3, "min": 4}
        agg_combo.setCurrentIndex(agg_map.get(ds.pivot.agg_func, 0))
        form.addRow("聚合方式:", agg_combo)
        filter_edit = LineEdit()
        filter_edit.setText(ds.pivot.filter_field or "")
        filter_edit.setPlaceholderText("可选")
        form.addRow("筛选字段:", filter_edit)
        filter_val = LineEdit()
        filter_val.setText(ds.pivot.filter_value or "")
        filter_val.setPlaceholderText("可选")
        form.addRow("筛选值:", filter_val)

        btns = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel)
        btns.accepted.connect(dlg.accept)
        btns.rejected.connect(dlg.reject)
        form.addRow(btns)

        if dlg.exec_():
            ds.sheet_name = sheet_combo.currentText()
            ds.is_pivot = pivot_check.isChecked()
            ds.pivot.row_field = row_edit.text().strip()
            ds.pivot.value_field = val_edit.text().strip()
            ds.pivot.agg_func = agg_combo.currentText().split()[0]
            ds.pivot.filter_field = filter_edit.text().strip()
            ds.pivot.filter_value = filter_val.text().strip()
            self._refresh_source_table()

    # ── Tab 4: 执行输出 ──
    def _init_run_tab(self):
        tab = QWidget()
        lay = QVBoxLayout(tab)
        lay.setSpacing(8)
        lay.addWidget(SubtitleLabel("4️⃣ 执行输出"))

        self._seed_info_label = CaptionLabel("种子值: 未提取（请在「字段配置」中为种子字段填写「数据源名!列名」）")
        lay.addWidget(self._seed_info_label)

        self._progress = ProgressBar()
        self._progress.setVisible(False)
        lay.addWidget(self._progress)

        btn_row = QHBoxLayout()
        refresh_btn = PushButton("🔄 刷新种子值")
        refresh_btn.clicked.connect(self._refresh_seed_info)
        btn_row.addWidget(refresh_btn)
        preview_btn = PushButton("👁️ 预览(前10条)")
        preview_btn.clicked.connect(self._preview)
        btn_row.addWidget(preview_btn)
        export_btn = PrimaryPushButton("💾 导出 Excel")
        export_btn.clicked.connect(self._export)
        btn_row.addWidget(export_btn)
        btn_row.addStretch()
        lay.addLayout(btn_row)

        self._result_view = TextEdit()
        self._result_view.setReadOnly(True)
        lay.addWidget(self._result_view, 1)

        self._tabs.addTab(tab, "🚀 执行输出")

    def _refresh_seed_info(self):
        vals, msg = self._extract_seed_values()
        self._seed_info_label.setText(f"🌱 {msg}")

    def _preview(self):
        if not self._fields or not self._template_path:
            InfoBar.warning("提示", "请先选择模板并配置字段", parent=self, duration=2000, position=InfoBarPosition.TOP)
            return
        vals, msg = self._extract_seed_values()
        self._seed_info_label.setText(f"🌱 {msg}")
        if not vals:
            InfoBar.warning("提示", msg, parent=self, duration=3000, position=InfoBarPosition.TOP)
            return
        try:
            self._load_all_sources()
            seed_field = self._get_seed_field_name()
            preview_vals = vals[:10]
            rows, errors = execute_pipeline(self._fields, self._data_sources, preview_vals, seed_field)
            self._show_results(rows, errors, is_preview=True)
        except Exception as e:
            InfoBar.error("执行失败", str(e), parent=self, duration=3000, position=InfoBarPosition.TOP)

    def _export(self):
        if not self._fields or not self._template_path:
            InfoBar.warning("提示", "请先选择模板", parent=self, duration=2000, position=InfoBarPosition.TOP)
            return
        vals, msg = self._extract_seed_values()
        self._seed_info_label.setText(f"🌱 {msg}")
        if not vals:
            InfoBar.warning("提示", msg, parent=self, duration=3000, position=InfoBarPosition.TOP)
            return
        path, _ = QFileDialog.getSaveFileName(self, "导出到", "output.xlsx", "Excel (*.xlsx)")
        if not path:
            return
        self._load_all_sources()
        seed_field = self._get_seed_field_name()
        self._progress.setVisible(True)
        self._progress.setValue(0)
        self._worker = PipelineWorker(self._fields, dict(self._data_sources), vals, seed_field)
        self._worker.progress.connect(lambda c, t: self._progress.setValue(int(c / t * 100)))
        self._worker.finished_ok.connect(lambda rows, errs: self._on_export_done(rows, errs, path))
        self._worker.error.connect(lambda e: self._on_export_error(e))
        self._worker.start()

    def _on_export_done(self, rows, errors, path):
        self._progress.setVisible(False)
        try:
            write_output(self._template_path, path, self._fields, rows)
            self._show_results(rows, errors, is_preview=False, path=path)
            InfoBar.success("导出成功", f"已保存到 {path}", parent=self, duration=3000, position=InfoBarPosition.TOP)
        except Exception as e:
            InfoBar.error("写入失败", str(e), parent=self, duration=3000, position=InfoBarPosition.TOP)

    def _on_export_error(self, msg):
        self._progress.setVisible(False)
        InfoBar.error("执行失败", msg, parent=self, duration=3000, position=InfoBarPosition.TOP)

    def _show_results(self, rows, errors, is_preview=True, path=None):
        fields = self._fields
        parts = []
        title = "👁️ 预览结果" if is_preview else f"✅ 导出完成: {path}"
        parts.append(f"<h3>{title}</h3>")
        parts.append(f"<p>共 {len(rows)} 行数据</p>")
        if errors:
            parts.append(f"<p style='color:#f44747'>⚠️ {len(errors)} 行有错误</p>")
        parts.append("<table border='1' cellpadding='4' style='border-collapse:collapse;'>")
        parts.append("<tr>" + "".join(f"<th>{f.name}</th>" for f in fields) + "</tr>")
        for r in rows:
            parts.append("<tr>" + "".join(f"<td>{r.get(f.name, '')}</td>" for f in fields) + "</tr>")
        parts.append("</table>")
        if errors:
            parts.append("<h4>错误详情:</h4><pre style='color:#f44747'>" + "\n".join(errors[:20]) + "</pre>")
        self._result_view.setHtml("".join(parts))

    # ── Tab 5: 帮助 ──
    def _init_help_tab(self):
        tab = QWidget()
        lay = QVBoxLayout(tab)
        view = TextEdit()
        view.setReadOnly(True)
        view.setHtml(self._help_html())
        lay.addWidget(view)
        self._tabs.addTab(tab, "📖 使用帮助")

    def _help_html(self):
        return '''
<h2>📊 Excel 数据处理 - 使用帮助</h2>

<h3>🔄 工作流程</h3>
<ol>
<li><b>选择输出模板</b> — 选一个 Excel 模板，插件读取第一行表头作为字段</li>
<li><b>配置字段取值</b> — 为每个字段设置取值公式，勾选种子字段</li>
<li><b>添加数据源</b> — 添加多个 Excel 数据源供公式查找/聚合</li>
<li><b>执行输出</b> — 预览/导出结果</li>
</ol>

<h3>🌱 种子字段配置</h3>
<p>种子字段决定输出行数，每个种子值生成一行。种子字段的「取值逻辑」填写<b>数据源名!列名</b>指定取值列。</p>
<pre>示例:
  种子字段「项目编码」的公式: 数据源A!项目编码
  → 从数据源A的「项目编码」列提取所有值作为种子

  种子字段「员工ID」的公式: 员工信息!A
  → 从数据源「员工信息」的A列提取所有值</pre>
<p>勾选顶部「🌱种子去重」可对提取的种子值去重（默认开启）。</p>

<h4>📦 种子提取器（可选）</h4>
<p>当取值列的单元格包含复合内容（如「编码 名称」），可在公式末尾加 <code>>> 提取器</code> 只取需要的部分：</p>
<pre>数据源A!项目库 >> SPLIT(" ", 0)
→ 取「项目库」列每行<b>空格前</b>的部分作为种子值

SPLIT(分隔符, 索引)
  分隔符: 用于分割的字符（如 " "、"\\", "-"）
  索引:    取分割后的第几段（从0开始）</pre>
<pre>示例原始值 → 提取结果（SPLIT(" ", 0)）:
  "5477FZ240001 基于历史覆冰..." → "5477FZ240001"
  "A-001 测试项目"              → "A-001"</pre>

<h3>📝 公式语法（非种子字段）</h3>
<p>用 <code>{字段名}</code> 引用<b>当前行</b>中<b>已计算的前序字段</b>。</p>

<h4>字段引用</h4>
<pre>{项目编码}              -- 引用前序字段"项目编码"的值
{合同金额} - {已付金额}  -- 算术运算
{项目名称} + " - " + {项目编码}  -- 文本拼接</pre>

<h4>查找函数</h4>
<pre>VLOOKUP({项目编码}, 数据源A!A:D, 2, FALSE)
-- 在数据源A中，用项目编码查找，返回第2列的值</pre>

<h4>条件聚合</h4>
<pre>SUMIF(数据源C!A:A, {项目编码}, 数据源C!D:D)
SUMIFS(数据源C!D:D, 数据源C!A:A, {项目编码}, 数据源C!B:B, "已付款")
COUNTIF(数据源C!A:A, {项目编码})
AVERAGEIF(数据源C!A:A, {项目编码}, 数据源C!D:D)
MAXIFS(数据源C!D:D, 数据源C!A:A, {项目编码})</pre>

<h4>自增函数 SEQ</h4>
<p>按行号自增，每行递增末尾数字段，保持原始位数（前导零补齐）。</p>
<pre>SEQ("1-1")     -- 1-1, 1-2, 1-3, 1-4...
SEQ(1)        -- 1, 2, 3, 4...
SEQ("A-001")  -- A-001, A-002, A-003...
SEQ("001")    -- 001, 002, 003...（保持3位补零）</pre>

<h4>其他函数</h4>
<pre>ROUND({金额}, 2)       -- 四舍五入
ABS({差额})            -- 绝对值
CONCATENATE({a}, "-", {b})  -- 拼接
IF({金额} > 0, "有", "无")  -- 条件判断</pre>

<h3>🔗 数据源引用格式</h3>
<p><code>数据源名!列范围</code>，例如：</p>
<ul>
<li><code>数据源A!A:D</code> — 数据源A的A到D列</li>
<li><code>数据源B!A:A</code> — 数据源B的A列</li>
</ul>

<h3>💡 典型场景</h3>
<pre>字段顺序:
  1. 项目编码  [☑种子] 公式: 数据源A!项目编码
  2. 项目名称  公式: VLOOKUP({项目编码}, 数据源A!A:D, 2, FALSE)
  3. 合同金额  公式: VLOOKUP({项目编码}, 数据源B!A:D, 3, FALSE)
  4. 已付金额  公式: SUMIF(数据源C!A:A, {项目编码}, 数据源C!D:D)
  5. 未付金额  公式: {合同金额} - {已付金额}
  6. 完成率    公式: ROUND({已付金额} / {合同金额} * 100, 1)</pre>

<p style='color:#888'>提示：所有输出单元格均为文本格式，避免数字精度丢失。</p>
'''

    def on_activate(self):
        super().on_activate()
